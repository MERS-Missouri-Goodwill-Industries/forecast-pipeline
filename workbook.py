"""Builds the 74-sheet planning workbook.

Everything is emitted as a live Excel FORMULA, never a computed value. The COO edits this
file after export, so a pre-computed number is silently wrong the moment he changes an
input. The reconciliation cell (Q6 on every store tab) exists to catch exactly that.
"""

from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from forecast_engine import (
    MONTH_NAMES,
    WEEKDAYS,
    build_day_factors,
    build_store_plan,
    compute_forecasted_bases,
    expand_closures,
    month_row_ranges,
)

# --- styling ---------------------------------------------------------------------------
INPUT_FONT = Font(name="Arial", size=10, color="FF0000FF")
INPUT_FILL = PatternFill("solid", fgColor="FFFFFF99")
FORMULA_FONT = Font(name="Arial", size=10, color="FF000000")
CROSS_FONT = Font(name="Arial", size=10, color="FF008000")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="FFCC0000")
CLOSURE_FILL = PatternFill("solid", fgColor="FFFFFF00")

EXEC_CAL_SHEET_NAME = "Exec Calendar Inputs"
CAL_OFFSET = 30  # Calendar Inputs content starts this many rows below the Exec Summary block

MONEY = "$#,##0"
MONEY2 = "$#,##0.00"
PCT4 = "0.0000%"

FIRST_DAILY_ROW = 9


def _hdr(cell, size: int = 10):
    cell.font = Font(name="Arial", size=size, bold=True)
    return cell


def _inp(cell, fmt: str | None = None):
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    return cell


def _fml(cell, fmt: str | None = None):
    cell.font = FORMULA_FONT
    if fmt:
        cell.number_format = fmt
    return cell


def _xs(cell, fmt: str | None = None):
    cell.font = CROSS_FONT
    if fmt:
        cell.number_format = fmt
    return cell


def _note(cell):
    cell.font = NOTE_FONT
    return cell


def sheet_ref(name: str, ref: str) -> str:
    """Quote a sheet name for use in a formula. Apostrophes must be doubled."""
    return f"'{name.replace(chr(39), chr(39) * 2)}'!{ref}"


def tab_name(store: dict) -> str:
    return f"{store['code']} - {store['name']}"[:31]


def _autofit_columns(ws, min_width: int = 8, max_width: int = 32, padding: int = 2) -> None:
    """Approximate Excel's AutoFit column width.

    openpyxl never renders formulas, so a formula cell's source text (a 70-character
    IF/SUMPRODUCT expression, say) is not what a reader sees on the page -- its number
    format is. Widths are estimated from that instead. Note-font cells are skipped: they
    are meant to overflow into the empty cells beside them, not to size the whole column
    around themselves.
    """
    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None or cell.font == NOTE_FONT:
                continue
            if isinstance(value, str) and value.startswith("="):
                fmt = cell.number_format or ""
                if "$" in fmt:
                    length = 14
                elif "%" in fmt:
                    length = 9
                elif fmt == "yyyy-mm-dd":
                    length = 11
                else:
                    length = 10
            elif isinstance(value, bool):
                length = 5
            elif isinstance(value, (int, float)):
                length = len(f"{value:,.0f}")
            elif hasattr(value, "isoformat"):
                length = 11
            else:
                length = len(str(value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)

    for col, length in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + padding))


def validate_plan(
    stores: list[dict],
    planned_sales: dict[str, float],
    days: list[dict],
    overrides: dict[str, dict],
    *,
    swing_multiplier: float = 3.0,
    tol: float = 0.01,
) -> list[dict]:
    """Sanity-check the plan before it ships. Pure Python -- no Excel involved, so this
    runs whether or not a formula would ever get the chance to recalculate.

    Returns a list of {"check", "status", "detail"} records, status one of
    "ERROR" | "FLAG" | "OK". build_workbook refuses to generate a file over an ERROR;
    a FLAG is a judgment call surfaced to the COO on the Data_Validation sheet, not a
    reason to block the build.
    """
    results: list[dict] = []
    codes = [s["code"] for s in stores]

    dupes = sorted({c for c in codes if codes.count(c) > 1})
    results.append({
        "check": "Store codes are unique",
        "status": "ERROR" if dupes else "OK",
        "detail": f"Duplicated: {dupes}" if dupes else f"All {len(codes)} codes unique.",
    })

    missing = [c for c in codes if c not in planned_sales]
    results.append({
        "check": "Every store has a Planned Sales figure",
        "status": "ERROR" if missing else "OK",
        "detail": f"Missing for: {missing}" if missing else "All stores accounted for.",
    })

    valid_statuses = {"Continuing", "New store", "Closed"}
    bad_status = sorted({s["code"] for s in stores if s["status"] not in valid_statuses})
    results.append({
        "check": "Store status is Continuing, New store, or Closed",
        "status": "ERROR" if bad_status else "OK",
        "detail": (f"Unrecognized status on: {bad_status}" if bad_status
                   else "All statuses recognized."),
    })

    negative = sorted(c for c, v in planned_sales.items() if v < 0)
    results.append({
        "check": "No store has a negative Planned Sales figure",
        "status": "ERROR" if negative else "OK",
        "detail": f"Negative on: {negative}" if negative else "No negative values.",
    })

    zero_continuing = sorted(
        s["code"] for s in stores
        if s["status"] == "Continuing" and abs(planned_sales.get(s["code"], 0.0)) < tol
    )
    results.append({
        "check": "No continuing store is committed to $0",
        "status": "FLAG" if zero_continuing else "OK",
        "detail": (f"$0 on: {zero_continuing} -- confirm this is a deliberate zero-out, not "
                   "a missed override." if zero_continuing else "No unexplained $0 stores."),
    })

    swings = []
    for s in stores:
        base = s.get("base_sales") or 0.0
        val = planned_sales.get(s["code"], 0.0)
        if base > 0 and val > 0:
            ratio = val / base
            if ratio > swing_multiplier or ratio < 1 / swing_multiplier:
                swings.append(f"{s['code']} ({ratio:.1f}x last year)")
    results.append({
        "check": f"No store's Planned Sales is more than {swing_multiplier:.0f}x off its "
                 "own prior-year Base Sales",
        "status": "FLAG" if swings else "OK",
        "detail": ("Review before committing: " + "; ".join(swings)) if swings
                   else "Every store is within a normal range of its own history.",
    })

    broken_recon = []
    for s in stores:
        val = planned_sales.get(s["code"], 0.0)
        plan = build_store_plan(val, days)
        if abs(plan["full_year_total"] - val) > tol:
            broken_recon.append(f"{s['code']} (${plan['full_year_total']:,.2f} vs ${val:,.2f})")
    results.append({
        "check": "Every store's day-of-week disaggregation adds back up to its own annual "
                 "Planned Sales",
        "status": "ERROR" if broken_recon else "OK",
        "detail": ("Broken on: " + "; ".join(broken_recon)) if broken_recon
                   else f"All {len(stores)} stores reconcile to the cent.",
    })

    network_total = sum(planned_sales.values())
    results.append({
        "check": "Network total is the sum of every store's Planned Sales",
        "status": "OK",
        "detail": f"${network_total:,.0f} across {len(stores)} stores.",
    })

    plan_year = days[0]["date_obj"].year if days else None
    reversed_ranges, outside_year = [], []
    for s in stores:
        for c in overrides.get(s["code"], {}).get("closures", []) or []:
            start = dt.date.fromisoformat(c["start"])
            end = dt.date.fromisoformat(c["end"])
            if start > end:
                reversed_ranges.append(f"{s['code']} ({c['start']} to {c['end']})")
            elif plan_year and (start.year != plan_year or end.year != plan_year):
                outside_year.append(f"{s['code']} ({c['start']} to {c['end']})")
    results.append({
        "check": "Every closure's start date is on or before its end date",
        "status": "ERROR" if reversed_ranges else "OK",
        "detail": ("Reversed range on: " + "; ".join(reversed_ranges)) if reversed_ranges
                   else "No reversed closure ranges.",
    })
    results.append({
        "check": f"Every closure falls within the plan year ({plan_year})",
        "status": "FLAG" if outside_year else "OK",
        "detail": (("Outside the plan year, so it has no effect: " + "; ".join(outside_year))
                   if outside_year else "No closures fall outside the plan year."),
    })

    return results


def build_workbook(
    *,
    year: int,
    stores: list[dict],
    weights: dict[str, float],
    holidays: list[dict],
    recommended_plan: float,
    store_overrides: dict[str, dict] | None = None,
    session_name: str | None = None,
    author: str = "vyamaykin@mersgoodwill.org",
) -> Workbook:
    session_name = session_name or f"POC Prototype {year} Planned Sales Workbook"
    overrides = store_overrides or {}
    days = build_day_factors(year, weights, holidays)
    ranges = month_row_ranges(days, FIRST_DAILY_ROW)
    daily_total_row = FIRST_DAILY_ROW + len(days)
    df_last = 1 + len(days)
    store_start = 8
    store_end = store_start + len(stores) - 1

    # No baseline forecast column ships in the workbook -- this split only seeds the one
    # COO Adjusted Planned Sales input per store, in Python, at build time. An explicit
    # override still wins.
    forecasted = compute_forecasted_bases(stores, recommended_plan)
    planned_sales: dict[str, float] = {}
    for st in stores:
        override_base = overrides.get(st["code"], {}).get("plan_base")
        planned_sales[st["code"]] = forecasted[st["code"]] if override_base is None else override_base

    validation_results = validate_plan(stores, planned_sales, days, overrides)
    hard_errors = [r for r in validation_results if r["status"] == "ERROR"]
    if hard_errors:
        raise ValueError(
            "Refusing to build a workbook over bad input data:\n"
            + "\n".join(f"- {r['check']}: {r['detail']}" for r in hard_errors)
        )

    wb = Workbook()
    wb.remove(wb.active)

    _how_to_use(wb, session_name, author)
    ecx = wb.create_sheet(EXEC_CAL_SHEET_NAME)
    dv = wb.create_sheet("Data_Validation")
    al = wb.create_sheet("All_Stores_Summary")
    mm = wb.create_sheet("Monthly_Matrix")
    dd = wb.create_sheet("Daily_Disaggregated_Plan")
    fg = wb.create_sheet("Final_Sales_Goals")
    _day_factors(wb, days, df_last)
    _plan_inputs(wb, stores, planned_sales, store_start, store_end)

    for idx, store in enumerate(stores):
        _store_tab(wb, store, store_start + idx, days, ranges, daily_total_row,
                   df_last, overrides.get(store["code"], {}), year)

    _exec_calendar_inputs(wb, ecx, session_name, stores, year, weights, holidays, df_last)
    _data_validation(wb, dv, stores, validation_results)
    _all_stores_summary(wb, al, stores, store_start)
    _monthly_matrix(wb, mm, stores)
    _daily_feed(wb, dd, stores, days)
    _final_sales_goals(wb, fg, stores, days)
    return wb


# --- How_To_Use ------------------------------------------------------------------------
def _how_to_use(wb: Workbook, session_name: str, author: str) -> None:
    s = wb.create_sheet("How_To_Use")
    s.sheet_properties.tabColor = "FFFD9D0D"
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 95

    _hdr(s["A1"], 14).value = f"{session_name} — How to Use This Workbook"

    _hdr(s["A3"]).value = "Purpose"
    s["A4"] = ("Turns each store's committed Planned Sales figure plus your day-of-week "
               "assumptions into a day-by-day, store-by-store sales plan. Everything below "
               "Exec Calendar Inputs and Plan_Inputs is a live formula — change an input and the "
               "rest of the workbook recalculates.")
    s["A4"].alignment = Alignment(wrap_text=True)

    _hdr(s["A6"]).value = "Colour Legend"
    legend = [
        ("Blue on Yellow", "Editable input. These are the only cells you type in.", _inp),
        ("Black", "Formula calculating within the same sheet.", _fml),
        ("Green", "Pulls a value from another sheet.", _xs),
        ("Red Italic", "Audit note explaining a rule. Not something to edit.", _note),
    ]
    for i, (label, desc, styler) in enumerate(legend):
        r = 7 + i
        s[f"A{r}"] = label
        styler(s[f"A{r}"])
        s[f"B{r}"] = desc

    _hdr(s["A12"]).value = "Sheet Index"
    index = [
        ("Exec Calendar Inputs", "Top-level numbers and the monthly rollup, plus your control "
         "centre for weekday weights, holidays, and Weekday Mix."),
        ("Data_Validation", "Sanity checks run once at generation, plus a live per-store "
         "reconciliation that recalculates as you edit."),
        ("All_Stores_Summary", "One row per store: COO Adjusted Planned Sales, code, name, region, status."),
        ("Monthly_Matrix", "Every store's 12 months in one grid, totalling to the network."),
        ("Daily_Disaggregated_Plan", "Flat store-by-day feed for Power BI."),
        ("Final_Sales_Goals", "Store code, plan date, plan value -- nothing else."),
        ("Day_Factors", "The 365-day engine behind every daily figure."),
        ("Plan_Inputs", "The plan total and each store's COO Adjusted Planned Sales input."),
        ("[Store] tabs", "One 365-day schedule per store. Closure dates live in columns V to X."),
    ]
    for i, (name, desc) in enumerate(index):
        r = 13 + i
        _fml(s[f"A{r}"]).value = name
        s[f"B{r}"] = desc

    _hdr(s["A23"]).value = "Reading a Store Tab"
    s["A24"] = ("Every store tab is laid out the same way: one 365-day daily schedule, one money "
                "column.")
    s["A24"].alignment = Alignment(wrap_text=True)

    guide = [
        ("Column E — Day % of Annual",
         "This day's slice of the whole year. The same seven values for every store, all year. "
         "A closed day reads 0%. See the Weekday Mix table on Exec Calendar Inputs."),
        ("Column D — COO Adjusted Plan",
         "What you are committing to for that day: the store's annual Planned Sales (S2) x "
         "Day % of Annual. This is the only money column on the sheet. Clear a cell to take "
         "that day out of the plan."),
        ("Column I — COO Month Total", "That month's COO Adjusted Plan days added up."),
        ("Row 3",
         "Your committed plan by month. It adds up the daily column, so clearing days moves it "
         "straight away."),
        ("Cell Q6",
         "The 365 daily amounts (Q3) minus the annual input (S2). Always $0.00 -- if it isn't, "
         "a formula has been broken."),
    ]
    for i, (label, desc) in enumerate(guide):
        r = 25 + i
        _hdr(s[f"A{r}"]).value = label
        s[f"B{r}"] = desc
        s[f"B{r}"].alignment = Alignment(wrap_text=True)

    _hdr(s["A34"]).value = "Taking days or months out"
    s["A35"] = ("Select any run of cells in the COO Adjusted Plan column and press Delete — a few "
                "days, half a month, or a whole month. The COO Month Total, that month in row 3, "
                "the year in Q3, and the Diff in Q6 all follow on their own. For a planned "
                "renovation, enter the date range in columns V to X instead and the days zero "
                "out automatically.")
    s["A35"].alignment = Alignment(wrap_text=True)

    _hdr(s["A38"]).value = "Contacts"
    s["A39"] = "Product Owner"
    s["B39"] = author


# --- Day_Factors -----------------------------------------------------------------------
def _day_factors(wb: Workbook, days: list[dict], df_last: int) -> None:
    s = wb.create_sheet("Day_Factors")
    headers = ["Date", "Month", "Day of Week", "Holiday", "Day Weight (% of Week)",
               "Day % of Annual", "Day % of Month", "Month Name"]
    widths = [14, 10, 14, 20, 20, 16, 16, 14]
    for i, (h, w) in enumerate(zip(headers, widths)):
        _hdr(s.cell(1, i + 1)).value = h
        s.column_dimensions[get_column_letter(i + 1)].width = w

    O = CAL_OFFSET
    cal_w = sheet_ref(EXEC_CAL_SHEET_NAME, f"$B${6 + O}:$B${12 + O}")
    cal_d = sheet_ref(EXEC_CAL_SHEET_NAME, f"$A${6 + O}:$A${12 + O}")
    cal_t = sheet_ref(EXEC_CAL_SHEET_NAME, f"$B${13 + O}")
    hol_lbl = sheet_ref(EXEC_CAL_SHEET_NAME, f"$B${16 + O}:$B${25 + O}")
    hol_dt = sheet_ref(EXEC_CAL_SHEET_NAME, f"$A${16 + O}:$A${25 + O}")

    for i, day in enumerate(days):
        r = 2 + i
        _inp(s[f"A{r}"]).value = day["date_obj"]
        s[f"A{r}"].number_format = "yyyy-mm-dd"
        _fml(s[f"B{r}"]).value = f"=MONTH(A{r})"
        _fml(s[f"C{r}"]).value = f'=TEXT(A{r},"dddd")'
        _xs(s[f"D{r}"]).value = f'=IFERROR(INDEX({hol_lbl}, MATCH(A{r}, {hol_dt}, 0)), "")'
        # Flat weekday share of a normal week -- same seven values all year.
        _xs(s[f"E{r}"], PCT4).value = f"=INDEX({cal_w}, MATCH(C{r}, {cal_d}, 0)) / {cal_t}"
        # Share of the WHOLE YEAR: weight over the combined weight of every open day.
        _fml(s[f"F{r}"], PCT4).value = (
            f'=IF(D{r}<>"", 0, E{r} / SUMPRODUCT(($D$2:$D${df_last}="")*($E$2:$E${df_last})))'
        )
        _fml(s[f"G{r}"], PCT4).value = (
            f"=IF(F{r}=0, 0, F{r}/SUMIF($B$2:$B${df_last}, B{r}, $F$2:$F${df_last}))"
        )
        _fml(s[f"H{r}"]).value = f'=TEXT(A{r},"mmmm")'

    tr = df_last + 1
    _hdr(s[f"A{tr}"]).value = "TOTAL"
    _fml(s[f"E{tr}"], PCT4).value = f"=SUM(E2:E{df_last})"
    _fml(s[f"F{tr}"], PCT4).value = f"=SUM(F2:F{df_last})"  # must read 100%

    s.column_dimensions["J"].width = 12
    s.column_dimensions["K"].width = 10
    _hdr(s["J1"]).value = "Weekday"
    _hdr(s["K1"]).value = "Count"
    for i, d in enumerate(WEEKDAYS):
        r = 2 + i
        _fml(s[f"J{r}"]).value = d
        _fml(s[f"K{r}"]).value = f"=COUNTIF($C$2:$C${df_last}, J{r})"
    _hdr(s["J9"]).value = "Total Days"
    _fml(s["K9"]).value = "=SUM(K2:K8)"
    s.freeze_panes = "A2"


# --- Plan_Inputs -----------------------------------------------------------------------
def _plan_inputs(wb: Workbook, stores: list[dict], planned_sales: dict[str, float],
                 start: int, end: int) -> None:
    s = wb.create_sheet("Plan_Inputs")
    s.column_dimensions["A"].width = 26
    s.column_dimensions["B"].width = 34
    for c in range(3, 6):
        s.column_dimensions[get_column_letter(c)].width = 22

    _hdr(s["A1"], 14).value = "Plan Inputs — Store Planned Sales"
    _hdr(s["A4"]).value = "Total Planned Sales"
    _fml(s["B4"], MONEY).value = f"=SUM($E${start}:$E${end})"

    header_row = start - 1
    for i, h in enumerate(["Code", "Name", "Region", "Status", "COO Adjusted Planned Sales"]):
        _hdr(s.cell(header_row, i + 1)).value = h
    s.cell(header_row, 5).comment = Comment(
        "Editable input. Seeded from a base-sales-weighted split of the network plan "
        "(new stores inherit the average of up to three same-region comparables; closed "
        "stores start at $0) -- change any value here to commit a different plan for that "
        "store.",
        "workbook.py",
    )

    for i, st in enumerate(stores):
        r = start + i
        _fml(s[f"A{r}"]).value = st["code"]
        _fml(s[f"B{r}"]).value = st["name"]
        _fml(s[f"C{r}"]).value = st["region"]
        _inp(s[f"D{r}"]).value = st["status"]
        _inp(s[f"E{r}"], MONEY).value = round(planned_sales[st["code"]], 2)

    tr = end + 1
    _hdr(s[f"A{tr}"]).value = "TOTAL"
    _fml(s[f"E{tr}"], MONEY).value = f"=SUM(E{start}:E{end})"


# --- store tab -------------------------------------------------------------------------
def _store_tab(wb: Workbook, store: dict, plan_row: int, days: list[dict],
               ranges: list[tuple[int, int]], daily_total_row: int, df_last: int,
               override: dict, year: int) -> None:
    s = wb.create_sheet(tab_name(store))
    month_cols = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]

    closures = override.get("closures", []) or []
    closed_dates = {c["date"] for c in expand_closures(closures)}
    has_closures = bool(closures)
    closures_end = 1 + max(1, len(closures))

    _hdr(s["A1"]).value = "Store Code"
    _hdr(s["B1"]).value = "Store Name"
    _hdr(s["D1"]).value = f"{year} Planned Sales"
    for m, name in enumerate(MONTH_NAMES):
        _hdr(s[f"{month_cols[m]}1"]).value = name[:3]
    _hdr(s["Q1"]).value = "Total"
    _hdr(s["S1"]).value = "COO Adjusted Planned Sales"
    _hdr(s["V1"]).value = "Closure Start"
    _hdr(s["W1"]).value = "Closure End"
    _hdr(s["X1"]).value = "Closure Label"

    _fml(s["A2"]).value = store["code"]
    _fml(s["B2"]).value = store["name"]
    _hdr(s["D3"]).value = "COO Adjusted Plan"

    for i in range(max(1, len(closures))):
        r = 2 + i
        if i < len(closures):
            _inp(s[f"V{r}"]).value = dt.date.fromisoformat(closures[i]["start"])
            s[f"V{r}"].number_format = "yyyy-mm-dd"
            _inp(s[f"W{r}"]).value = dt.date.fromisoformat(closures[i]["end"])
            s[f"W{r}"].number_format = "yyyy-mm-dd"
            _inp(s[f"X{r}"]).value = closures[i].get("label") or "Store Closure"
        else:
            _inp(s[f"V{r}"]); _inp(s[f"W{r}"]); _inp(s[f"X{r}"])

    for m in range(12):
        a, b = ranges[m]
        # Row 3 ADDS UP the daily COO column. This is what lets whole months or a handful
        # of days be cleared in the daily block and have the month, year and Q6 check follow.
        _xs(s[f"{month_cols[m]}3"], MONEY).value = f"=SUM(D{a}:D{b})"

    _fml(s["Q3"], MONEY).value = f"=D{daily_total_row}"
    _xs(s["S2"], MONEY).value = f'={sheet_ref("Plan_Inputs", f"$E${plan_row}")}'
    # Daily disaggregation (Q3) minus the annual input (S2) -- zero by construction. Anything
    # else means a formula somewhere has been broken.
    _fml(s["Q6"], MONEY2).value = "=Q3-$S$2"

    _hdr(s["A7"]).value = "Daily Schedule"

    headers = {"A": "Date", "B": "Day of Week", "D": "COO Adjusted Plan",
               "E": "Day % of Annual", "G": "Month",
               "I": "COO Month Total"}
    for col, h in headers.items():
        _hdr(s[f"{col}8"]).value = h

    month_start = FIRST_DAILY_ROW
    for i, day in enumerate(days):
        r = FIRST_DAILY_ROW + i
        dfr = 2 + i
        _xs(s[f"A{r}"]).value = f'={sheet_ref("Day_Factors", f"A{dfr}")}'
        s[f"A{r}"].number_format = "yyyy-mm-dd"
        _xs(s[f"B{r}"]).value = f'={sheet_ref("Day_Factors", f"C{dfr}")}'
        _inp(s[f"D{r}"], MONEY2).value = f"=$S$2 * E{r}"

        if has_closures:
            hol = sheet_ref("Day_Factors", f"D{dfr}")
            _xs(s[f"F{r}"]).value = (
                f'=IF({hol}<>"", {hol}, IF(SUMPRODUCT(($V$2:$V${closures_end}<=A{r})'
                f'*($W$2:$W${closures_end}>=A{r}))>0, "Store Closure", ""))'
            )
            _fml(s[f"E{r}"], PCT4).value = (
                f'=IF(F{r}<>"", 0, {sheet_ref("Day_Factors", f"E{dfr}")} / '
                f'SUMPRODUCT(($F${FIRST_DAILY_ROW}:$F${daily_total_row - 1}="")'
                f'*({sheet_ref("Day_Factors", f"$E$2:$E${df_last}")})))'
            )
        else:
            _xs(s[f"E{r}"], PCT4).value = f'={sheet_ref("Day_Factors", f"F{dfr}")}'

        _fml(s[f"G{r}"]).value = f'=TEXT(A{r},"mmmm")'

        last_of_month = (i == len(days) - 1) or (days[i + 1]["month"] != day["month"])
        if last_of_month:
            _fml(s[f"I{r}"], MONEY).value = f"=SUM(D{month_start}:D{r})"
            month_start = r + 1

        if day["date"] in closed_dates:
            for c in "ABDE":
                s[f"{c}{r}"].fill = CLOSURE_FILL

    _hdr(s[f"A{daily_total_row}"]).value = "FULL YEAR"
    for col, fmt in [("D", MONEY), ("E", PCT4)]:
        _fml(s[f"{col}{daily_total_row}"], fmt).value = (
            f"=SUM({col}{FIRST_DAILY_ROW}:{col}{daily_total_row - 1})"
        )
    _note(s[f"A{daily_total_row + 2}"]).value = (
        "Yellow cells with blue text are yours to change — the COO Adjusted Plan column. "
        "Monthly totals add up the daily rows, so clearing days moves the month and the year."
    )
    _autofit_columns(s)
    s.freeze_panes = "A9"


# --- rollups ---------------------------------------------------------------------------
def _exec_calendar_inputs(wb: Workbook, s, session_name: str, stores: list[dict], year: int,
                          weights: dict[str, float], holidays: list[dict], df_last: int) -> None:
    s.sheet_properties.tabColor = "FF0065A4"
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 20
    s.column_dimensions["C"].width = 60
    for c in range(4, 8):
        s.column_dimensions[get_column_letter(c)].width = 18

    # ---- Executive Summary ----
    _hdr(s["A1"], 14).value = f"{session_name} — Executive Summary"

    rows = [
        ("Plan Year", year, None, None),
        ("COO Adjusted Planned Sales ($)", f'={sheet_ref("Plan_Inputs", "$B$4")}', MONEY,
         "The committed network plan -- the sum of every store's Planned Sales on Plan_Inputs."),
        ("Continuing Stores", sum(1 for x in stores if x["status"] == "Continuing"), None, None),
        ("New Stores", sum(1 for x in stores if x["status"] == "New store"), None, None),
        ("Total Stores", len(stores), None, None),
    ]
    for i, (label, val, fmt, note) in enumerate(rows):
        r = 4 + i
        _hdr(s[f"A{r}"]).value = label
        cell = s[f"B{r}"]
        cell.value = val
        (_xs if isinstance(val, str) and str(val).startswith("=") else _inp if val is None else _fml)(cell, fmt)
        if note:
            _note(s[f"C{r}"]).value = note

    _hdr(s["A14"]).value = "Monthly Breakdown"
    for i, h in enumerate(["Month", "% of Annual", "Planned Sales ($)"]):
        _hdr(s.cell(15, i + 1)).value = h
    O = CAL_OFFSET
    for m in range(12):
        r = 16 + m
        _fml(s[f"A{r}"]).value = MONTH_NAMES[m]
        # Same sheet now -- the month table below (rows 29+O onward) holds this directly.
        _fml(s[f"B{r}"], PCT4).value = f"=E{30 + O + m}"
        _xs(s[f"C{r}"], MONEY).value = (
            f'={sheet_ref("Monthly_Matrix", f"${get_column_letter(3 + m)}${3 + len(stores)}")}'
        )

    # ---- Calendar Inputs — Control Centre ----
    _hdr(s[f"A{1 + O}"], 14).value = "Calendar Inputs — Control Centre"
    _note(s[f"A{2 + O}"]).value = ("One weight column. The seven values are each day's share of a "
                                   "normal week and total 100%. Everything else here is calculated.")

    _hdr(s[f"A{5 + O}"]).value = "Weekday"
    _hdr(s[f"B{5 + O}"]).value = "Normalized % of Week"
    total = sum(weights.get(d, 0.0) for d in WEEKDAYS) or 1.0
    normalized = [weights.get(d, 0.0) / total for d in WEEKDAYS]
    # Round each independently and the sum can drift off 100% by a hair (e.g. 99.9999%).
    # Round all but the last, then let the last absorb whatever residual is left, so the
    # seven inputs always sum to exactly 1.0 -- not just approximately.
    rounded = [round(v, 6) for v in normalized[:-1]]
    rounded.append(round(1.0 - sum(rounded), 6))
    for i, d in enumerate(WEEKDAYS):
        r = 6 + i + O
        _fml(s[f"A{r}"]).value = d
        _inp(s[f"B{r}"], PCT4).value = rounded[i]
    _hdr(s[f"A{13 + O}"]).value = "Total"
    _fml(s[f"B{13 + O}"], PCT4).value = f"=SUM(B{6 + O}:B{12 + O})"

    _hdr(s[f"A{15 + O}"]).value = "Holiday Date"
    _hdr(s[f"B{15 + O}"]).value = "Holiday Label"
    _hdr(s[f"C{15 + O}"]).value = "Day of Week"
    for i in range(10):
        r = 16 + i + O
        if i < len(holidays):
            _inp(s[f"A{r}"]).value = dt.date.fromisoformat(holidays[i]["date"])
            s[f"A{r}"].number_format = "yyyy-mm-dd"
            _inp(s[f"B{r}"]).value = holidays[i]["label"]
        else:
            _inp(s[f"A{r}"])
            _inp(s[f"B{r}"])
        _fml(s[f"C{r}"]).value = f'=IF(A{r}="","",TEXT(A{r},"dddd"))'

    for i, label in enumerate(["Month", "Days", "Closed Days", "Selling Days",
                               "% of Annual", "Month Name"]):
        _hdr(s.cell(28 + O, i + 1)).value = label
    for m in range(1, 13):
        r = 29 + m + O
        _fml(s[f"A{r}"]).value = m
        _xs(s[f"B{r}"]).value = f'=COUNTIF({sheet_ref("Day_Factors", f"$B$2:$B${df_last}")}, A{r})'
        _xs(s[f"C{r}"]).value = (
            f'=SUMPRODUCT(({sheet_ref("Day_Factors", f"$B$2:$B${df_last}")}=A{r})'
            f'*(LEN({sheet_ref("Day_Factors", f"$D$2:$D${df_last}")})>0))'
        )
        _fml(s[f"D{r}"]).value = f"=B{r}-C{r}"
        _xs(s[f"E{r}"], PCT4).value = (
            f'=SUMIF({sheet_ref("Day_Factors", f"$B$2:$B${df_last}")}, A{r}, '
            f'{sheet_ref("Day_Factors", f"$F$2:$F${df_last}")})'
        )
        _fml(s[f"F{r}"]).value = f'=TEXT(DATE(2000,A{r},1),"mmmm")'

    _note(s[f"A{44 + O}"]).value = ('Closed-day counts use SUMPRODUCT+LEN>0, not COUNTIFS "<>", '
                                    'because Excel treats a formula returning "" as non-blank.')

    # Weekday Mix
    _hdr(s[f"A{46 + O}"], 12).value = "Weekday Mix — Normalized % of Week"
    for i, label in enumerate(["Day", "Wkdy #", "Day % of Annual", "% of Week"]):
        _hdr(s.cell(47 + O, i + 1)).value = label
    wk_num = {"Sunday": 1, "Monday": 2, "Tuesday": 3, "Wednesday": 4,
              "Thursday": 5, "Friday": 6, "Saturday": 7}
    for i, d in enumerate(WEEKDAYS):
        r = 48 + i + O
        _fml(s[f"A{r}"]).value = d
        _fml(s[f"B{r}"]).value = wk_num[d]
        _xs(s[f"C{r}"], PCT4).value = (
            f'=IFERROR(SUMIF({sheet_ref("Day_Factors", f"$C$2:$C${df_last}")}, A{r}, '
            f'{sheet_ref("Day_Factors", f"$F$2:$F${df_last}")})'
            f' / COUNTIFS({sheet_ref("Day_Factors", f"$C$2:$C${df_last}")}, A{r}, '
            f'{sheet_ref("Day_Factors", f"$D$2:$D${df_last}")}, ""), 0)'
        )
        _fml(s[f"D{r}"], "0.0%").value = f"=IFERROR(C{r}/$C${55 + O}, 0)"
    _hdr(s[f"A{55 + O}"]).value = "Total"
    _fml(s[f"C{55 + O}"], PCT4).value = f"=SUM(C{48 + O}:C{54 + O})"
    _fml(s[f"D{55 + O}"], "0.0%").value = f"=SUM(D{48 + O}:D{54 + O})"
    _hdr(s[f"A{56 + O}"]).value = "Weekend (Sat+Sun)"
    _fml(s[f"D{56 + O}"], "0.0%").value = f"=D{53 + O}+D{54 + O}"
    _hdr(s[f"A{57 + O}"]).value = "Weekday (Mon-Fri)"
    _fml(s[f"D{57 + O}"], "0.0%").value = f"=SUM(D{48 + O}:D{52 + O})"
    _note(s[f"A{59 + O}"]).value = ("These seven day-shares are the same for every store and the "
                                    "same all year: an open Friday is worth the same slice of the "
                                    "year in January as in July. Day % of Annual totals 100% "
                                    "across all days.")


def _all_stores_summary(wb: Workbook, s, stores: list[dict], start: int) -> None:
    headers = ["Store Code", "Store Name", "Region", "Status", "COO Adjusted Planned Sales ($)"]
    for i, h in enumerate(headers):
        _hdr(s.cell(1, i + 1)).value = h
    s.column_dimensions["A"].width = 12
    s.column_dimensions["B"].width = 32
    s.column_dimensions["C"].width = 12
    s.column_dimensions["D"].width = 14
    s.column_dimensions["E"].width = 26

    for i, st in enumerate(stores):
        r, pr = 2 + i, start + i
        _fml(s[f"A{r}"]).value = st["code"]
        _fml(s[f"B{r}"]).value = st["name"]
        _fml(s[f"C{r}"]).value = st["region"]
        _fml(s[f"D{r}"]).value = st["status"]
        _xs(s[f"E{r}"], MONEY).value = f'={sheet_ref("Plan_Inputs", f"$E${pr}")}'

    tr = 2 + len(stores)
    _hdr(s[f"A{tr}"]).value = "NETWORK TOTAL"
    _fml(s[f"E{tr}"], MONEY).value = f"=SUM(E2:E{tr - 1})"
    s.freeze_panes = "A2"


def _monthly_matrix(wb: Workbook, s, stores: list[dict]) -> None:
    cols = [get_column_letter(3 + m) for m in range(12)]
    _hdr(s["A2"]).value = "Code"
    _hdr(s["B2"]).value = "Store Name"
    for m, name in enumerate(MONTH_NAMES):
        _hdr(s[f"{cols[m]}2"]).value = name
    _hdr(s["O2"]).value = "Total"
    s.column_dimensions["A"].width = 12
    s.column_dimensions["B"].width = 32
    for c in cols + ["O"]:
        s.column_dimensions[c].width = 14

    src_cols = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
    for i, st in enumerate(stores):
        r = 3 + i
        name = tab_name(st)
        _fml(s[f"A{r}"]).value = st["code"]
        _fml(s[f"B{r}"]).value = st["name"]
        for m in range(12):
            # Row 3 on the store tab = the COO Adjusted track.
            _xs(s[f"{cols[m]}{r}"], MONEY).value = f"={sheet_ref(name, f'${src_cols[m]}$3')}"
        _xs(s[f"O{r}"], MONEY).value = f"={sheet_ref(name, '$Q$3')}"

    tr = 3 + len(stores)
    _hdr(s[f"A{tr}"]).value = "NETWORK TOTAL"
    for c in cols + ["O"]:
        _fml(s[f"{c}{tr}"], MONEY).value = f"=SUM({c}3:{c}{tr - 1})"
    s.freeze_panes = "C3"


def _daily_feed(wb: Workbook, s, stores: list[dict], days: list[dict]) -> None:
    headers = ["Store Code", "Store Name", "Date", "Month", "Day of Week",
               "Day % of Annual", "COO Adjusted Plan ($)"]
    widths = [12, 32, 14, 12, 14, 16, 22]
    for i, (h, w) in enumerate(zip(headers, widths)):
        _hdr(s.cell(1, i + 1)).value = h
        s.column_dimensions[get_column_letter(i + 1)].width = w

    r = 2
    for st in stores:
        name = tab_name(st)
        for i in range(len(days)):
            sr = FIRST_DAILY_ROW + i
            _fml(s.cell(r, 1)).value = st["code"]
            _fml(s.cell(r, 2)).value = st["name"]
            c3 = s.cell(r, 3); _xs(c3).value = f"={sheet_ref(name, f'$A${sr}')}"
            c3.number_format = "yyyy-mm-dd"
            _fml(s.cell(r, 4)).value = f'=TEXT(C{r},"mmmm")'
            _xs(s.cell(r, 5)).value = f"={sheet_ref(name, f'$B${sr}')}"
            _xs(s.cell(r, 6), PCT4).value = f"={sheet_ref(name, f'$E${sr}')}"
            _xs(s.cell(r, 7), MONEY2).value = f"={sheet_ref(name, f'$D${sr}')}"
            r += 1
    s.freeze_panes = "A2"


def _final_sales_goals(wb: Workbook, s, stores: list[dict], days: list[dict]) -> None:
    """The bare export: store_code, plan_date, plan_value -- nothing else."""
    headers = ["store_code", "plan_date", "plan_value"]
    widths = [12, 14, 18]
    for i, (h, w) in enumerate(zip(headers, widths)):
        _hdr(s.cell(1, i + 1)).value = h
        s.column_dimensions[get_column_letter(i + 1)].width = w

    r = 2
    for st in stores:
        name = tab_name(st)
        for i in range(len(days)):
            sr = FIRST_DAILY_ROW + i
            _fml(s.cell(r, 1)).value = st["code"]
            c2 = s.cell(r, 2); _xs(c2).value = f"={sheet_ref(name, f'$A${sr}')}"
            c2.number_format = "yyyy-mm-dd"
            _xs(s.cell(r, 3), MONEY2).value = f"={sheet_ref(name, f'$D${sr}')}"
            r += 1
    s.freeze_panes = "A2"


_VALIDATION_FILL = {
    "OK": PatternFill("solid", fgColor="FFC6EFCE"),
    "FLAG": PatternFill("solid", fgColor="FFFFEB9C"),
    "ERROR": PatternFill("solid", fgColor="FFFFC7CE"),
}


def _data_validation(wb: Workbook, s, stores: list[dict], results: list[dict]) -> None:
    """Two tables: what validate_plan checked once in Python at generation time (static --
    editing the file afterward does not change these rows), and a live per-store
    reconciliation that recalculates every time the file is opened or edited."""
    s.sheet_properties.tabColor = "FFCC0000"
    s.column_dimensions["A"].width = 55
    s.column_dimensions["B"].width = 10
    s.column_dimensions["C"].width = 70
    s.column_dimensions["D"].width = 12

    _hdr(s["A1"], 14).value = "Data Validation"
    _note(s["A2"]).value = (
        "Checked at Generation (below) ran once in Python when this workbook was built -- "
        "it will NOT change if you edit a cell. Live Reconciliation, further down, "
        "recalculates every time you open or edit this file."
    )

    _hdr(s["A4"]).value = "Checked at Generation"
    for i, h in enumerate(["Check", "Status", "Detail"]):
        _hdr(s.cell(5, i + 1)).value = h
    for i, r in enumerate(results):
        row = 6 + i
        s[f"A{row}"].value = r["check"]
        cell = s[f"B{row}"]
        cell.value = r["status"]
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = _VALIDATION_FILL.get(r["status"], PatternFill())
        s[f"C{row}"].value = r["detail"]
        s[f"C{row}"].alignment = Alignment(wrap_text=True)

    live_title_row = 7 + len(results)
    _hdr(s[f"A{live_title_row}"], 12).value = "Live Reconciliation — recalculates on open"
    header_row = live_title_row + 1
    for i, h in enumerate(["Code", "Store Name", "Q6 (should be $0.00)", "Status"]):
        _hdr(s.cell(header_row, i + 1)).value = h

    first_data_row = header_row + 1
    for i, st in enumerate(stores):
        row = first_data_row + i
        name = tab_name(st)
        _fml(s[f"A{row}"]).value = st["code"]
        _fml(s[f"B{row}"]).value = st["name"]
        _xs(s[f"C{row}"], MONEY2).value = f"={sheet_ref(name, '$Q$6')}"
        _fml(s[f"D{row}"]).value = f'=IF(ABS(C{row})<0.01,"OK","CHECK")'

    last_data_row = first_data_row + len(stores) - 1
    total_row = last_data_row + 1
    _hdr(s[f"A{total_row}"]).value = "Any store failing?"
    _fml(s[f"D{total_row}"]).value = (
        f'=IF(COUNTIF(D{first_data_row}:D{last_data_row},"CHECK")=0,"OK","CHECK")'
    )
    s.freeze_panes = "A6"


def workbook_bytes(**kwargs) -> bytes:
    buf = BytesIO()
    build_workbook(**kwargs).save(buf)
    return buf.getvalue()